#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync_f259.py — ซิงก์ทะเบียนกฎหมาย F-259 (รอบที่ 1 ปี 2569) เข้าสู่แอป LexGuard

อ่าน:  lexguard/local-data/F-259_2569_R1.xlsx  (gitignore แล้ว — ห้าม commit)
เขียน: (ก) supabase/seed-data.json          — โครงเดิม {laws, catname, meta}
       (ข) supabase/migrations/016_sync_f259_2569r1.sql — ชุด upsert สำหรับฐานข้อมูลจริง

หลักการสำคัญ:
  * ดึงข้อมูลจาก Excel จริงเท่านั้น ไม่มีการแต่งเนื้อหา
  * ตรวจคอลัมน์จากหัวตาราง (keyword) เพราะแต่ละชีทวางคอลัมน์ไม่ตรงกัน (LF เลื่อน 1 คอลัมน์, CCS ไม่มีคอลัมน์รหัส)
  * รหัสซ้ำข้ามหมวดคงไว้ตามเอกสารจริง (LG ใช้ LF-xxx ภายใต้ cat='LG') — พึ่ง unique (cat, code)
  * migration ไม่ทำลายข้อมูลผู้ใช้: อัปเดตเฉพาะฟิลด์ทะเบียนของ lg_laws และ "insert ข้อกำหนดเฉพาะกฎหมายที่ยังไม่มีข้อกำหนดในฐาน"
"""
import openpyxl, json, re, sys, os, datetime
from collections import OrderedDict, Counter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                      # lexguard/
XLSX = os.path.join(ROOT, 'local-data', 'F-259_2569_R1.xlsx')
SEED_OUT = os.path.join(ROOT, 'supabase', 'seed-data.json')
SQL_OUT  = os.path.join(ROOT, 'supabase', 'migrations', '017_sync_f259_2569r1.sql')

# cat -> ชื่อชีท
SHEETS = OrderedDict([
    ('LA', 'LA-บริหารจัดการความปลอดภัย'),
    ('LB', 'LB-ไฟฟ้าและพลังงาน'),
    ('LC', 'LC-อัคคีภัย (OK)'),
    ('LD', 'LD-ร้อน แสง เสียง สภาพแวดล้อม'),
    ('LE', 'LE-ก่อสร้าง ลิฟต์  เครื่องจักร '),
    ('LF', 'LF- Service'),
    ('LG', 'LG-คณะกรรมการสวัสดิการ.'),
    ('CC', 'FG - CCS'),
])
CATNAME = OrderedDict([
    ('LA', 'การบริหารจัดการความปลอดภัย อาชีวอนามัยฯ'),
    ('LB', 'ไฟฟ้าและพลังงาน'),
    ('LC', 'การป้องกันและระงับอัคคีภัย'),
    ('LD', 'ความร้อน แสงสว่าง เสียง สภาพแวดล้อม'),
    ('LE', 'ก่อสร้าง ลิฟต์ เครื่องจักร ปั้นจั่น'),
    ('LF', 'Service'),
    ('LG', 'คณะกรรมการสวัสดิการ'),
    ('CC', 'CCS (คุ้มครองข้อมูลส่วนบุคคล/ไซเบอร์)'),
])
CC_NOTE = 'รหัสสร้างโดยระบบ — เอกสาร F-259 ยังไม่กำหนดรหัสหมวดนี้'
CODE_RE = re.compile(r'^[A-Za-z]{1,3}[- ]?\d{1,3}')

def cell(ws, r, c):
    return ws.cell(r, c).value if c else None

def s(v):
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return v.strftime('%d/%m/%Y')
    return str(v).strip()

def find_col(ws, kw, rows=(5, 6, 7)):
    for r in rows:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and kw in str(v):
                return c
    return None

def marked(v):
    """เครื่องหมาย C/NC: รองรับ '/', 'P', 'ü' (Wingdings), boolean True — ไม่นับ False/ว่าง/-/x"""
    if v is None or v is False:
        return False
    t = str(v).strip()
    if t == '' or t.lower() in ('false', '0', 'no', '-', 'x', 'na', 'n/a'):
        return False
    return True

def detect(ws):
    m = find_col(ws, 'กระทรวง')
    return {
        'ministry': m,
        'name':   find_col(ws, 'ชื่อกฎหมาย'),
        'req':    find_col(ws, 'สรุปสาระสำคัญ'),
        'date':   find_col(ws, 'วันที่ประกาศ'),
        'resp':   find_col(ws, 'ผู้รับผิดชอบ'),
        'freq':   find_col(ws, 'ความถี่'),
        'report': find_col(ws, 'การรายงานผล'),
        'docs':   find_col(ws, 'เอกสารที่เกี่ยวข้อง'),
        'note':   find_col(ws, 'หมายเหตุ'),
        'cmark':  next((c for c in range(1, ws.max_column + 1) if s(ws.cell(7, c).value) == 'C'), None),
        'ncmark': next((c for c in range(1, ws.max_column + 1) if s(ws.cell(7, c).value) == 'NC'), None),
        'code':   1 if (m and m > 1) else None,   # coded sheets: code=A ; CCS: no code col
    }

def parse_sheet(ws, cat):
    """คืนลิสต์ law dict: {cat, code|None, ministry, name, date, status, repealed, reqs[]}"""
    col = detect(ws)
    laws = []
    cur = None
    repealed = False
    for r in range(8, ws.max_row + 1):
        codev = cell(ws, r, col['code'])
        namev = cell(ws, r, col['name'])
        code_s = s(codev)
        name_s = s(namev)
        # เข้าสู่ส่วน "กฎหมายที่ยกเลิก" — หัวข้อในคอลัมน์รหัส: "กฎหมายที่ยกเลิก" (LA/LB) หรือ "ยกเลิก" เดี่ยวๆ (LD)
        # (ไม่รวมคำว่า 'ยกเลิก' ที่อยู่ในเนื้อหาคอลัมน์อื่น)
        if col['code'] and (code_s.startswith('กฎหมายที่ยกเลิก') or code_s == 'ยกเลิก'):
            repealed = True
            cur = None
            continue
        # ตรวจจุดเริ่มกฎหมายใหม่
        if col['code']:
            has_code = bool(codev) and bool(CODE_RE.match(code_s))
            has_name = bool(name_s)
            if has_code and has_name:
                is_start, new_code = True, code_s
            elif not has_code and has_name:
                # กฎหมายที่ไม่มีรหัสในชีท (เช่น LG ประกาศกรมฯ / กฎหมายยกเลิกบางฉบับ) — สร้างรหัสภายหลัง
                is_start, new_code = True, None
            else:
                # มีรหัสแต่ "ไม่มีชื่อ" = แถว placeholder ว่าง (เช่น รหัส LG ที่ยังไม่เติมข้อมูล) → ข้าม
                is_start, new_code = False, None
        else:  # CCS — เริ่มกฎหมายใหม่เมื่อมีชื่อกฎหมาย (คอลัมน์ B); แถวหัวข้อ full-width จะไม่มีชื่อ
            is_start = bool(name_s)
            new_code = None
        if is_start:
            cur = {
                'cat': cat,
                'code': new_code,
                'ministry': s(cell(ws, r, col['ministry'])),
                'name': name_s,
                'date': s(cell(ws, r, col['date'])),
                'repealed': repealed,
                'reqs': [],
            }
            laws.append(cur)
        # เก็บข้อกำหนดของแถวนี้ (ถ้ามีเนื้อหาใน "สรุปสาระสำคัญ")
        if cur is not None:
            req_t = s(cell(ws, r, col['req']))
            if req_t:
                c_ok = marked(cell(ws, r, col['cmark']))
                nc_ok = marked(cell(ws, r, col['ncmark']))
                status = 'unmet' if (nc_ok and not c_ok) else 'met'
                cur['reqs'].append({
                    't': req_t,
                    's': status,
                    'resp': s(cell(ws, r, col['resp'])),
                    'freq': s(cell(ws, r, col['freq'])),
                    'docs': s(cell(ws, r, col['docs'])),
                    'report': s(cell(ws, r, col['report'])),
                    'note': s(cell(ws, r, col['note'])),
                })
    return laws

def summary_nc_notes(wb):
    """ดึงหมายเหตุยาวของรายการ NC จากชีทสรุปความสอดคล้อง (ระบุ LA-031 / LA-032 พร้อมเหตุผล 'รอประกาศหลักสูตร')
    คืน dict {code: note_text}"""
    notes = {}
    for sheet in wb.sheetnames:
        if 'สรุป' not in sheet:
            continue
        ws = wb[sheet]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not v or ('LA-031' not in str(v) and 'LA-032' not in str(v)):
                    continue
                for m in re.finditer(r'(L[A-Z]-\d{3})\s*:\s*(.*?)(?=(?:L[A-Z]-\d{3}\s*:)|$)', str(v), re.S):
                    body = m.group(2).strip()
                    if body:
                        notes[m.group(1).strip()] = body
    return notes


def main():
    if not os.path.exists(XLSX):
        sys.exit('ไม่พบไฟล์ Excel: ' + XLSX)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    all_laws = []
    for cat, sheet in SHEETS.items():
        all_laws.extend(parse_sheet(wb[sheet], cat))

    # ── กำหนดรหัสสำหรับ CCS (สร้างโดยระบบ) ─────────────────────────────
    cc_seq = 0
    for law in all_laws:
        if law['cat'] == 'CC':
            cc_seq += 1
            law['code'] = 'CC-%03d' % cc_seq
            law['sys_note'] = CC_NOTE

    # ── จัดการรหัสซ้ำในส่วน "ยกเลิก" ที่ชนกับฉบับ active หมวดเดียวกัน ──────
    active_codes = {}
    for law in all_laws:
        if not law['repealed'] and law['code']:
            active_codes.setdefault(law['cat'], set()).add(law['code'])
    rep_counter = Counter()
    for law in all_laws:
        if not law['repealed']:
            continue
        if law['code'] and law['code'] in active_codes.get(law['cat'], set()):
            # เก็บ code เดิมไว้ในหมายเหตุ แล้วต่อท้าย -R เพื่อไม่ชน unique(cat, code)
            law['sys_note'] = 'รหัสเดิมในทะเบียน: %s (ซ้ำกับฉบับปัจจุบัน)' % law['code']
            law['code'] = law['code'] + '-R'
        elif not law['code']:
            rep_counter[law['cat']] += 1
            law['code'] = '%s-R%02d' % (law['cat'], rep_counter[law['cat']])
            law['sys_note'] = 'กฎหมายที่ยกเลิก (ไม่มีรหัสในเอกสารต้นฉบับ) — รหัสสร้างโดยระบบ'

    # ── หมวด LG: ใช้รหัส LG-nnn ตรงกับ production เดิม (ชีทใส่รหัส LF-xxx cross-list ที่ชน LF) ──
    # นำเข้าเฉพาะฉบับที่มีเนื้อหา (แถว placeholder ว่างถูกข้ามตั้งแต่ parse แล้ว)
    lg_seq = 0
    for law in all_laws:
        if law['cat'] == 'LG' and not law['repealed']:
            lg_seq += 1
            law['code'] = 'LG-%03d' % lg_seq

    # ── กฎหมาย active ที่ไม่มีรหัสในชีท (เช่น LD พรบ.เงินทดแทน) → generate รหัสต่อท้ายเลขสูงสุดของหมวด ──
    maxnum = {}
    for law in all_laws:
        m = re.search(r'(\d+)$', law['code']) if law['code'] else None
        if m:
            maxnum[law['cat']] = max(maxnum.get(law['cat'], 0), int(m.group(1)))
    for law in all_laws:
        if not law['code']:
            n = maxnum.get(law['cat'], 0) + 1
            maxnum[law['cat']] = n
            law['code'] = '%s-%03d' % (law['cat'], n)
            law['sys_note'] = 'รหัสสร้างโดยระบบ — เอกสาร F-259 ไม่ได้ระบุรหัสสำหรับฉบับนี้'

    # ── ตรวจความปลอดภัย: ทุกฉบับต้องมี code และ name ไม่ว่าง ──────────────
    bad = [(l['cat'], l.get('code'), l['name'][:30]) for l in all_laws if not l.get('code') or not (l.get('name') or '').strip()]
    if bad:
        sys.exit('พบกฎหมายที่ code/name ว่าง (แก้ parser ก่อน): ' + repr(bad[:10]))

    # ── ผูกสถานะกฎหมาย + เหตุผลยกเลิก ─────────────────────────────────
    for law in all_laws:
        if law['repealed']:
            law['status'] = 'repealed'
            # เหตุผล: ใช้ note ของข้อกำหนดแรกที่มี ไม่งั้นใช้ข้อความมาตรฐาน
            reason = next((q['note'] for q in law['reqs'] if q['note']), '')
            law['repeal_reason'] = reason or 'ยกเลิกตามทะเบียน F-259 (ส่วนกฎหมายที่ยกเลิก)'
        else:
            any_unmet = any(q['s'] == 'unmet' for q in law['reqs'])
            law['status'] = 'bad' if any_unmet else 'ok'

    # ── แนบหมายเหตุยาวจากชีทสรุปเข้ากับข้อกำหนด NC (LA-031 / LA-032) ────
    long_notes = summary_nc_notes(wb)
    for law in all_laws:
        extra = long_notes.get(law['code'])
        if not extra:
            continue
        for q in law['reqs']:
            if q['s'] == 'unmet':
                q['note'] = (q['note'] + ' | ' + extra) if q['note'] else extra

    # ── สร้าง seed-data.json (โครงเดิม) ───────────────────────────────
    seed_laws = []
    for law in all_laws:
        obj = OrderedDict()
        obj['cat'] = law['cat']
        obj['code'] = law['code']
        obj['ministry'] = law['ministry']
        obj['name'] = law['name']
        obj['date'] = law['date']
        obj['reqs'] = [OrderedDict([('t', q['t']), ('s', q['s']), ('resp', q['resp']),
                                    ('freq', q['freq']), ('docs', q['docs']),
                                    ('report', q['report']), ('note', q['note'])]) for q in law['reqs']]
        obj['status'] = law['status']
        if law.get('sys_note'):
            obj['note'] = law['sys_note']
        if law['status'] == 'repealed':
            obj['repeal_reason'] = law.get('repeal_reason', '')
        seed_laws.append(obj)

    meta = OrderedDict([
        ('company', 'บริษัท จัสเทล เน็ทเวิร์ค'),
        ('doc', 'F-259 Rev.1'),
        ('cycle', 'รอบที่ 1 (ม.ค.-มี.ค.) ปี 2569'),
        ('updated', '26 ธันวาคม 2568'),
    ])
    seed = OrderedDict([('laws', seed_laws), ('catname', CATNAME), ('meta', meta)])
    with open(SEED_OUT, 'w', encoding='utf-8') as f:
        json.dump(seed, f, ensure_ascii=False, indent=1)

    # ── สร้าง migration 016 (upsert ที่ไม่ทำลายข้อมูลผู้ใช้) ──────────────
    write_sql(all_laws)

    # ── รายงานสรุป ────────────────────────────────────────────────────
    report(all_laws)

def sqlstr(v):
    if v is None or v == '':
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def write_sql(all_laws):
    out = []
    W = out.append
    W("-- 017_sync_f259_2569r1.sql — ซิงก์ทะเบียน F-259 รอบที่ 1 ปี 2569 (สร้างอัตโนมัติโดย scripts/sync_f259.py)")
    W("-- ปลอดภัยต่อข้อมูลผู้ใช้:")
    W("--   • lg_laws: upsert on conflict (cat, code) — อัปเดตเฉพาะฟิลด์ทะเบียน (ชื่อ/กระทรวง/วันที่/สถานะ)")
    W("--     ไม่แตะ review_date, active, หรือฟิลด์อื่นที่ผู้ใช้แก้ในแอป")
    W("--   • lg_requirements: insert เฉพาะกฎหมายที่ 'ยังไม่มีข้อกำหนดในฐาน' เท่านั้น")
    W("--     (กฎหมายเดิมที่ผู้ใช้ประเมิน C/NC หรือแนบหลักฐานไว้แล้ว จะไม่ถูกเขียนทับ)")
    W("--   • ไม่แตะ lg_attachments / lg_process_tracker / lg_requirements.evaluated_* / evidence")
    W("")
    W("begin;")
    W("")
    W("-- (0) หมวดใหม่ CCS")
    W("insert into lg_categories (code, name, color, sort_order) values")
    W("  ('CC', 'CCS (คุ้มครองข้อมูลส่วนบุคคล/ไซเบอร์)', '#00b3a4', 80)")
    W("on conflict (code) do update set name = excluded.name;")
    W("")
    W("-- (1) upsert กฎหมายทั้งหมด")
    for law in all_laws:
        note = law.get('sys_note', '')
        if law['status'] == 'repealed':
            W("insert into lg_laws (cat, code, ministry, name, issue_date, status, repeal_reason) values")
            W("  (%s, %s, %s, %s, %s, 'repealed', %s)" % (
                sqlstr(law['cat']), sqlstr(law['code']), sqlstr(law['ministry']),
                sqlstr(law['name']), sqlstr(law['date']), sqlstr(law.get('repeal_reason', ''))))
            W("on conflict (cat, code) do update set")
            W("  ministry = excluded.ministry, name = excluded.name, issue_date = excluded.issue_date,")
            W("  status = 'repealed', repeal_reason = excluded.repeal_reason, updated_at = now();")
        else:
            W("insert into lg_laws (cat, code, ministry, name, issue_date, status) values")
            W("  (%s, %s, %s, %s, %s, %s)" % (
                sqlstr(law['cat']), sqlstr(law['code']), sqlstr(law['ministry']),
                sqlstr(law['name']), sqlstr(law['date']), sqlstr(law['status'])))
            W("on conflict (cat, code) do update set")
            W("  ministry = excluded.ministry, name = excluded.name,")
            W("  issue_date = excluded.issue_date, status = excluded.status, updated_at = now();")
    W("")
    W("-- (2) insert ข้อกำหนด เฉพาะกฎหมายที่ยังไม่มีข้อกำหนดในฐาน (กันเขียนทับงานผู้ใช้)")
    for law in all_laws:
        if not law['reqs']:
            continue
        law_key = "(select id from lg_laws where cat=%s and code=%s)" % (sqlstr(law['cat']), sqlstr(law['code']))
        W("insert into lg_requirements (law_id, seq, text, status, responsible, frequency, documents, note)")
        W("select lid, v.seq, v.text, v.status, v.responsible, v.frequency, v.documents, v.note from")
        W("  (select id as lid from lg_laws where cat=%s and code=%s) L," % (sqlstr(law['cat']), sqlstr(law['code'])))
        W("  (values")
        rows = []
        for i, q in enumerate(law['reqs']):
            # การรายงานผล (report) รวมเข้ากับ documents เพื่อเก็บให้ครบ (lg_requirements ไม่มีคอลัมน์แยก)
            docs = q['docs']
            if q['report']:
                docs = (docs + '\nการรายงานผล: ' + q['report']).strip('\n')
            rows.append("    (%d, %s, %s, %s, %s, %s, %s)" % (
                i, sqlstr(q['t']), sqlstr('met' if q['s'] == 'met' else 'unmet'),
                sqlstr(q['resp']), sqlstr(q['freq']), sqlstr(docs), sqlstr(q['note'])))
        W(",\n".join(rows))
        W("  ) as v(seq, text, status, responsible, frequency, documents, note)")
        W("where not exists (select 1 from lg_requirements r where r.law_id = L.lid);")
        W("")
    W("commit;")
    with open(SQL_OUT, 'w', encoding='utf-8') as f:
        f.write("\n".join(out) + "\n")

def report(all_laws):
    print("=" * 66)
    print(" รายงานซิงก์ F-259 รอบที่ 1 ปี 2569")
    print("=" * 66)
    by_cat = OrderedDict()
    for cat in list(SHEETS.keys()):
        rows = [l for l in all_laws if l['cat'] == cat]
        act = [l for l in rows if l['status'] != 'repealed']
        rep = [l for l in rows if l['status'] == 'repealed']
        bad = [l for l in act if l['status'] == 'bad']
        by_cat[cat] = (len(act), len(rep), len(bad))
        note = ''
        if rep:
            note = '  ยกเลิก: ' + ', '.join(l['code'] for l in rep)
        print(" %-3s | active %3d | repealed %2d | NC(active) %d%s" %
              (cat, len(act), len(rep), len(bad), note))
    print("-" * 66)
    total_act = sum(v[0] for v in by_cat.values())
    total_rep = sum(v[1] for v in by_cat.values())
    total_nc = sum(len([q for q in l['reqs'] if q['s'] == 'unmet']) for l in all_laws)
    print(" รวม active = %d | repealed = %d | ข้อกำหนด NC = %d" % (total_act, total_rep, total_nc))
    print(" เป้าหมาย: LA 42, LB 16, LC 11, LD 14, LE 10, LF 55, LG 32, CC ~40")
    # NC detail
    print("-" * 66)
    print(" รายการ NC (ข้อกำหนดที่ยังไม่สอดคล้อง):")
    for l in all_laws:
        for q in l['reqs']:
            if q['s'] == 'unmet':
                print("   %s-%s | %s | note: %s" % (l['cat'], l['code'], q['t'][:40], (q['note'] or '(ว่าง)')[:50]))
    print("=" * 66)
    print(" เขียนไฟล์:")
    print("   " + os.path.relpath(SEED_OUT, ROOT))
    print("   " + os.path.relpath(SQL_OUT, ROOT))

if __name__ == '__main__':
    main()
