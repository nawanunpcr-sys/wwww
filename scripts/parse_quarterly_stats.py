import re
import openpyxl

PATH = 'source-data/F-259_ทะเบียนกฎหมายและติดตามความสอดคล้อง.xlsx'

def find_year(text):
    m = re.search(r'ปี\s*\.*\s*(\d{4})', text or '')
    return int(m.group(1)) if m else None

def parse_sheet(ws):
    # header rows are at row 4 (groups) and row 5 (quarter labels); data starts row 6
    year = find_year(ws.cell(row=3, column=1).value)
    if year is None:
        return None
    rows = []
    for r in range(6, ws.max_row + 1):
        code = ws.cell(row=r, column=1).value
        if not code or code in ('รวม',):
            continue
        rep = [ws.cell(row=r, column=c).value or 0 for c in (4, 5, 6, 7)]
        new = [ws.cell(row=r, column=c).value or 0 for c in (8, 9, 10, 11)]
        rows.append({'cat': code, 'repealed_q': rep, 'added_q': new})
    return {'year_be': year, 'rows': rows}

def main():
    wb = openpyxl.load_workbook(PATH, data_only=True)
    by_year = {}  # year_be -> (sheet_name, parsed)  keep last occurrence
    for name in wb.sheetnames:
        if 'Masterlist SHE Law' not in name and 'สรุป' not in name:
            continue
        ws = wb[name]
        parsed = parse_sheet(ws)
        if not parsed:
            print(f'-- skip sheet (no year found): {name}')
            continue
        by_year[parsed['year_be']] = (name, parsed)

    for year_be in sorted(by_year):
        name, parsed = by_year[year_be]
        print(f'=== ปี พ.ศ. {year_be} (sheet: {name}) ===')
        for row in parsed['rows']:
            print(f"  {row['cat']:>4}  ยกเลิก(Q1-4)={row['repealed_q']}  ใหม่(Q1-4)={row['added_q']}")
        tot_rep = [sum(r['repealed_q'][i] for r in parsed['rows']) for i in range(4)]
        tot_new = [sum(r['added_q'][i] for r in parsed['rows']) for i in range(4)]
        print(f"  รวมยกเลิก(Q1-4)={tot_rep}  รวมใหม่(Q1-4)={tot_new}")
        print()

if __name__ == '__main__':
    main()
